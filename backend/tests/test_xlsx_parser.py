import pytest
from openpyxl import Workbook

from aurum_encuestas.errors import XlsxParseError
from aurum_encuestas.xlsx_parser import parse_xlsx, _detect_breakdowns


def test_parse_detects_sample_size(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    assert db.sample_size == 500


def test_parse_detects_breakdowns(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    ids = [b.id for b in db.breakdowns]
    assert "general" in ids
    assert "sexo" in ids
    assert "edad" in ids
    assert "nse" in ids
    assert "punto" in ids


def test_parse_breakdown_sexo_categories(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    sexo = next(b for b in db.breakdowns if b.id == "sexo")
    assert "Hombre" in sexo.categories
    assert "Mujer" in sexo.categories


def test_parse_invalid_file_raises(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not an xlsx")
    with pytest.raises(XlsxParseError):
        parse_xlsx(str(bad))


def test_parse_detects_questions(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    assert len(db.questions) >= 1
    q1 = db.questions[0]
    assert q1.code == "P1"
    assert q1.text  # non-empty
    assert "Sí" in q1.options
    assert "No" in q1.options


def test_parse_question_confidence(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    assert q1.confidence >= 0.9  # $pN.label marker = high confidence


def test_parse_detects_three_column_blocks(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    blocks = db.data_blocks
    assert "counts_cols" in blocks
    assert "pct_row_cols" in blocks
    # counts block starts at col 3
    assert blocks["counts_cols"][0] == 3
    # second block detected
    assert blocks["pct_row_cols"][0] >= 19


def _ws_with_custom_breakdown(tmp_path):
    wb = Workbook()
    ws = wb.active
    # Row 1 headers: a known one (Sexo) + an unknown one (Religión)
    ws.cell(1, 4, "Sexo")
    ws.cell(1, 6, "Religión")
    # Row 2 sub-categories; General anchors block 1 at col 3
    ws.cell(2, 3, "General")
    ws.cell(2, 4, "Hombre")
    ws.cell(2, 5, "Mujer")
    ws.cell(2, 6, "Católico")
    ws.cell(2, 7, "Evangélico")
    out = tmp_path / "custom_bd.xlsx"
    wb.save(out)
    from openpyxl import load_workbook
    return load_workbook(out, data_only=True).worksheets[0]


def test_detect_breakdowns_includes_unknown_header(tmp_path):
    ws = _ws_with_custom_breakdown(tmp_path)
    bds = _detect_breakdowns(ws)
    by_label = {b.label: b for b in bds}
    # Unknown header is no longer dropped
    assert "Religión" in by_label
    assert by_label["Religión"].id == "religión"
    assert by_label["Religión"].categories == ["Católico", "Evangélico"]
    # Known header keeps its canonical id via the alias map
    assert by_label["Sexo"].id == "sexo"


def test_parse_detects_total_row(valid_xlsx_path):
    from aurum_encuestas.xlsx_parser import parse_xlsx
    db = parse_xlsx(str(valid_xlsx_path))
    # Fixture writes "Total" at row 3, col B
    assert db.total_row == 3
