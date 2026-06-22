from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule


def _make_source(n_options=2, bds_spec=None):
    q = SimpleNamespace(options=[f"opt{i}" for i in range(n_options)])
    all_bds = {}
    for bd_id, label, cats in (bds_spec or []):
        all_bds[bd_id] = {
            "label": label,
            "categories": {cat: opts for cat, opts in cats},
        }
    return SimpleNamespace(
        question=q,
        all_breakdowns_data=all_bds,
        breakdown_ids=[bd_id for bd_id, _, _ in (bds_spec or [])],
    )


def test_single_panel_has_own_label_col():
    """Each bd has its own label col at its left; group_header merge spans
    label_col + cat cols."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.92, "count": 230}, "opt1": {"pct": 0.08, "count": 20}}),
            ("40-59", {"opt0": {"pct": 0.91, "count": 228}, "opt1": {"pct": 0.09, "count": 22}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    # Group header merge spans DATA cols only (excludes label col A): B2:C2
    assert any(str(mr) == "B2:C2" for mr in ws.merged_cells.ranges), \
        f"expected B2:C2 merge, got {[str(m) for m in ws.merged_cells.ranges]}"
    assert ws["B2"].value == "Edad"
    assert ws["A2"].value is None
    # Label col A: cat row empty, counts row "Observaciones", option rows = labels
    assert ws["A3"].value is None
    assert ws["A4"].value == "Observaciones"
    assert ws["A5"].value == "opt0"
    assert ws["A6"].value == "opt1"
    # Cat headers in row 3 at cols B, C
    assert ws["B3"].value == "18-39"
    assert ws["C3"].value == "40-59"
    # Counts row 4 — sum option counts per cat
    assert ws["B4"].value == 250
    assert ws["C4"].value == 250
    # Option row 5: opt0 pct
    assert abs(ws["B5"].value - 0.92) < 1e-9
    assert abs(ws["C5"].value - 0.91) < 1e-9


def test_multi_panel_n_independent_tables():
    """N=2 bds → 2 independent group_header merges, separator col empty."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
            ("40-59", {"opt0": {"pct": 0.8, "count": 80}, "opt1": {"pct": 0.2, "count": 20}}),
        ]),
        ("sexo", "Sexo", [
            ("F", {"opt0": {"pct": 0.7, "count": 70}, "opt1": {"pct": 0.3, "count": 30}}),
            ("M", {"opt0": {"pct": 0.85, "count": 85}, "opt1": {"pct": 0.15, "count": 15}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad", "sexo"])
    wb = load_workbook(buf)
    ws = wb.active
    # First bd Edad: label col A + data B-C → merge B2:C2.
    # Second bd Sexo: NO own label col (shares first's leyenda visually) →
    # data starts at E directly, merge E2:F2.
    merge_strs = {str(m) for m in ws.merged_cells.ranges}
    assert "B2:C2" in merge_strs
    assert "E2:F2" in merge_strs
    assert ws["B2"].value == "Edad"
    assert ws["E2"].value == "Sexo"
    # Spacer col D row 2 empty
    assert ws["D2"].value is None
    assert ws["D3"].value is None
    assert ws["D4"].value is None
    # Only FIRST bd has its label col with Observaciones.
    assert ws["A4"].value == "Observaciones"
    # Second bd has NO label col rendered (col E is bd2's first data col).
    assert ws["E4"].value != "Observaciones"


def test_databar_per_bd_panel_scoped():
    """DataBarRule scoped to one bd's data cols, not entire sheet."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
        ("sexo", "Sexo", [
            ("F", {"opt0": {"pct": 0.7, "count": 70}, "opt1": {"pct": 0.3, "count": 30}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad", "sexo"])
    wb = load_workbook(buf)
    ws = wb.active
    # Collect databar rule ranges
    bar_ranges = []
    for cf_range, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if getattr(rule, "dataBar", None) is not None:
                bar_ranges.append(str(cf_range.sqref))
    import re
    assert bar_ranges, "expected at least one databar rule"
    # First bd Edad has label A + data B (1 cat). Spacer C.
    # Second bd Sexo has NO label col → data D (1 cat). Spacer E.
    # Each databar range scoped to ONE col (single-cat per bd).
    edad_pattern = re.compile(r"^B\d+(:B\d+)?$")
    sexo_pattern = re.compile(r"^D\d+(:D\d+)?$")
    for r in bar_ranges:
        # Must not span spacer cols C or E.
        assert ":C" not in r and ":E" not in r, f"databar range {r} spans into spacer col"
        assert edad_pattern.match(r) or sexo_pattern.match(r), \
            f"databar range {r} does not match expected per-bd pattern (Edad=B or Sexo=D)"


def test_hex_palette_applied():
    """Header cells use hex fill 999999 + font FFFFFF; body cells use FFFFFF + 000000."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    # Group header at B2 (data col, not label col A) fill check
    gh = ws["B2"]
    assert "999999" in (gh.fill.fgColor.value or "").upper()
    assert "FFFFFF" in (gh.font.color.rgb or "").upper()
    # Body option-row cell (B5) fill+font check
    body = ws["B5"]
    assert "FFFFFF" in (body.fill.fgColor.value or "").upper()
    assert "000000" in (body.font.color.rgb or "").upper()


def test_gridlines_hidden():
    """Worksheet must hide gridlines (showGridLines = False)."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.sheet_view.showGridLines is False, "worksheet gridlines must be hidden"


def test_cell_borders_applied():
    """Data + header cells have visible borders (thin BFBFBF)."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    # Header B2 (single cat → first+last): top + left + right, NO bottom.
    h = ws["B2"]
    assert h.border.top.style == "thin"
    assert h.border.left.style == "thin"
    assert h.border.right.style == "thin"
    assert h.border.bottom is None or h.border.bottom.style is None
    # Option row B5 first option (j=0, n_opts=2 → not last): single-col so first+last.
    # Has: right_always, left (first col), no bottom (not last opt).
    b5 = ws["B5"]
    assert b5.border.right.style == "thin"
    assert b5.border.left.style == "thin"
    assert b5.border.top is None or b5.border.top.style is None
    assert b5.border.bottom is None or b5.border.bottom.style is None
    # Last option row B6: same + bottom (outer).
    b6 = ws["B6"]
    assert b6.border.bottom.style == "thin"
    assert b6.border.right.style == "thin"


def test_empty_breakdown_groups_returns_empty_xlsx():
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[])
    buf = build_xlsx_for_table(src, [])
    wb = load_workbook(buf)
    ws = wb.active
    # No merged ranges at row 2
    assert not any(str(mr).startswith(("A2", "B2", "C2")) for mr in ws.merged_cells.ranges)


def test_xlsx_show_legend_true_includes_label_col():
    """show_legend=True → label col with 'Observaciones' + option labels."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    assert ws["A4"].value == "Observaciones"
    assert ws["A5"].value == "opt0"


def test_xlsx_show_legend_false_skips_label_col():
    """show_legend=False → no label col; data starts at col A."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
            ("40-59", {"opt0": {"pct": 0.8, "count": 80}, "opt1": {"pct": 0.2, "count": 20}}),
        ]),
    ])
    src.show_legend = False
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    # Cat headers at A3 + B3 (no label col), data at A5 + B5
    assert ws["A3"].value == "18-39"
    assert ws["B3"].value == "40-59"
    assert abs(ws["A5"].value - 0.9) < 1e-9
    # No "Observaciones" text anywhere (label col was skipped)
    assert ws["A4"].value != "Observaciones"


def test_databar_color_is_d9d9d9():
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    for cf_range, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            db = getattr(rule, "dataBar", None)
            if db is not None and db.color is not None:
                assert "D9D9D9" in (db.color.value or "").upper()
                return
    raise AssertionError("no DataBarRule found")


def test_data_cell_alignment_left():
    """Data cells text left-aligned with indent (matches Fase R design target)."""
    from aurum_encuestas.element_renderers.xlsx_builder import build_xlsx_for_table

    src = _make_source(n_options=2, bds_spec=[
        ("edad", "Edad", [
            ("18-39", {"opt0": {"pct": 0.9, "count": 90}, "opt1": {"pct": 0.1, "count": 10}}),
        ]),
    ])
    src.show_legend = True
    buf = build_xlsx_for_table(src, ["edad"])
    wb = load_workbook(buf)
    ws = wb.active
    cell = ws["B5"]
    assert cell.alignment.horizontal == "left"
    assert cell.alignment.indent == 1
