from openpyxl import Workbook, load_workbook
from aurum_encuestas.data_extractor import extract_chart_data, _resolve_breakdown_cols
from aurum_encuestas.xlsx_parser import parse_xlsx


def test_extract_chart_data_general(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks)
    assert "Total" in data
    # data is {breakdown_category: {option: {count, pct}}}
    assert data["Total"]["Sí"]["count"] == 458


def test_extract_chart_data_sexo(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    data = extract_chart_data(str(valid_xlsx_path), q1, "sexo", db.data_blocks)
    assert "Hombre" in data
    assert "Mujer" in data


def _ws_custom(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 6, "Religión")        # unknown header at col 6
    ws.cell(2, 3, "General")
    ws.cell(2, 6, "Católico")
    ws.cell(2, 7, "Evangélico")
    out = tmp_path / "res.xlsx"
    wb.save(out)
    return load_workbook(out, data_only=True).worksheets[0]


def test_resolve_breakdown_cols_generic(tmp_path):
    ws = _ws_custom(tmp_path)
    cols = _resolve_breakdown_cols(ws, "religión", block_start_col=3)
    assert cols == {"Católico": 6, "Evangélico": 7}


def test_extract_chart_data_allowed_categories(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    data = extract_chart_data(
        str(valid_xlsx_path), q1, "sexo", db.data_blocks,
        allowed_categories=["Mujer"],
    )
    assert list(data.keys()) == ["Mujer"]  # Hombre filtered out


def test_extract_chart_data_computes_pct(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks, total_row=db.total_row)
    assert data["Total"]["Sí"]["count"] == 458
    assert abs(data["Total"]["Sí"]["pct"] - 458 / 500) < 1e-9


def test_extract_chart_data_no_total_row_pct_none(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks, total_row=None)
    assert data["Total"]["Sí"]["pct"] is None
    assert data["Total"]["Sí"]["count"] == 458


def test_extract_chart_data_override_count(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    key = f"{q1.id}|general|Total|Sí"
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks,
                              total_row=db.total_row, overrides={key: {"count": 999}})
    assert data["Total"]["Sí"]["count"] == 999
    assert abs(data["Total"]["Sí"]["pct"] - 458 / 500) < 1e-9  # pct untouched


def test_extract_chart_data_override_pct(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    key = f"{q1.id}|general|Total|Sí"
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks,
                              total_row=db.total_row, overrides={key: {"pct": 0.5}})
    assert data["Total"]["Sí"]["pct"] == 0.5
    assert data["Total"]["Sí"]["count"] == 458  # count untouched


def test_count_cells_filter_excludes_unmarked(valid_xlsx_path):
    from aurum_encuestas.data_extractor import _find_question_rows, _resolve_breakdown_cols
    from openpyxl import load_workbook
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    ws = load_workbook(str(valid_xlsx_path), data_only=True).worksheets[0]
    rows = _find_question_rows(ws, q1)
    cols = _resolve_breakdown_cols(ws, "general", db.data_blocks["counts_cols"][0])
    si_cell = [rows["Sí"], cols["Total"]]          # mark ONLY "Sí"
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks,
                              total_row=db.total_row, count_cells=[si_cell])
    assert data["Total"]["Sí"]["count"] == 458     # marked → read
    assert data["Total"]["No"]["count"] == 0       # unmarked → 0
    assert data["Total"]["No"]["pct"] == 0.0


def test_count_cells_empty_reads_all(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    full = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks,
                              total_row=db.total_row, count_cells=None)
    none = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks,
                              total_row=db.total_row)
    assert full == none                            # None ⇒ no filter
    assert full["Total"]["No"]["count"] != 0       # "No" still read


def test_count_cells_override_wins(valid_xlsx_path):
    db = parse_xlsx(str(valid_xlsx_path))
    q1 = db.questions[0]
    key = f"{q1.id}|general|Total|No"
    # "No" is excluded by count_cells (empty marked set for it) but overridden to 5
    data = extract_chart_data(str(valid_xlsx_path), q1, "general", db.data_blocks,
                              total_row=db.total_row, count_cells=[[1, 1]],
                              overrides={key: {"count": 5}})
    assert data["Total"]["No"]["count"] == 5       # override applies after the 0-forcing
