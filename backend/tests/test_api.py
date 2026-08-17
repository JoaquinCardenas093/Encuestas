from unittest.mock import patch

from fastapi.testclient import TestClient

from aurum_encuestas.api import app

client = TestClient(app)


def test_health():
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json() == {"status": "ok"}


def test_parse_xlsx_endpoint(valid_xlsx_path):
    with open(valid_xlsx_path, "rb") as f:
        r = client.post("/api/parse-xlsx", files={"file": ("v.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    body = r.json()
    assert body["sample_size"] == 500
    assert any(b["id"] == "sexo" for b in body["breakdowns"])


def test_parse_xlsx_invalid_returns_400(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"junk")
    with open(bad, "rb") as f:
        r = client.post("/api/parse-xlsx", files={"file": ("bad.xlsx", f, "application/octet-stream")})
    assert r.status_code == 400
    assert r.json()["code"] == "xlsx_parse_error"


def test_parse_template_endpoint(valid_template_path):
    with open(valid_template_path, "rb") as f:
        r = client.post("/api/parse-template", files={"file": ("t.pptx", f, "application/vnd.openxmlformats-officedocument.presentationml.presentation")})
    assert r.status_code == 200
    body = r.json()
    assert body["shell_slide_index"] == 0
    assert "@Titulo" in body["placeholders"]


def test_parse_template_invalid_returns_400(invalid_template_one_slide):
    with open(invalid_template_one_slide, "rb") as f:
        r = client.post("/api/parse-template", files={"file": ("t.pptx", f, "application/octet-stream")})
    assert r.status_code == 400
    assert r.json()["code"] == "template_invalid"


def test_save_load_project_endpoints(tmp_path):
    proj = {
        "version": 1,
        "project_name": "Test",
        "inputs": {"db_path": "./x.xlsx", "template_path": "./t.pptx", "font_override": None},
        "slides": [],
    }
    r = client.post("/api/save-project", json={"name": "p", "state": proj}, headers={"X-Session-Id": "sess-a"})
    assert r.status_code == 200
    r2 = client.post("/api/load-project", json={"name": "p"}, headers={"X-Session-Id": "sess-a"})
    assert r2.status_code == 200
    assert r2.json()["project_name"] == "Test"


def test_save_load_project_by_name_isolated(tmp_path, monkeypatch):
    monkeypatch.setenv("HOME", str(tmp_path))
    state = {"project_name": "P1", "inputs": {"db_path": "x", "template_path": "y"}}
    s = client.post("/api/save-project", json={"name": "p1", "state": state}, headers={"X-Session-Id": "s1"})
    assert s.status_code == 200 and s.json()["name"] == "p1"
    l1 = client.post("/api/load-project", json={"name": "p1"}, headers={"X-Session-Id": "s1"})
    assert l1.status_code == 200 and l1.json()["project_name"] == "P1"
    l2 = client.post("/api/load-project", json={"name": "p1"}, headers={"X-Session-Id": "s2"})
    assert l2.status_code >= 400


def test_save_project_name_traversal_is_contained(tmp_path, monkeypatch):
    monkeypatch.setenv("HOME", str(tmp_path))
    state = {"project_name": "Evil", "inputs": {"db_path": "x", "template_path": "y"}}
    client.post("/api/save-project", json={"name": "../../evil", "state": state}, headers={"X-Session-Id": "s3"})
    projects_dir = tmp_path / ".aurum" / "sessions" / "s3" / "projects"
    written = list(projects_dir.glob("*.aurum.json"))
    assert len(written) == 1
    assert written[0].name == "evil.aurum.json"


class TestPreviewSlide:
    """Tests for /api/preview-slide endpoint"""

    def test_preview_slide_returns_base64_png(self, valid_template_path, valid_xlsx_path):
        """preview-slide endpoint returns base64-encoded PNG if LibreOffice available."""
        import shutil
        if not shutil.which("soffice"):
            import pytest
            pytest.skip("LibreOffice not installed")

        state_dict = {
            "version": 1,
            "project_name": "Test",
            "inputs": {"db_path": str(valid_xlsx_path), "template_path": str(valid_template_path), "font_override": None},
            "slides": [
                {
                    "id": "slide_0",
                    "type": "separator",
                    "title": "Test Slide",
                }
            ],
        }

        req = {
            "state": state_dict,
            "slide_index": 0,
        }
        r = client.post("/api/preview-slide", json=req)
        assert r.status_code == 200
        body = r.json()
        assert "png_base64" in body
        assert isinstance(body["png_base64"], str)
        assert len(body["png_base64"]) > 0

    def test_preview_slide_returns_placeholder_without_libreoffice(self, valid_template_path, valid_xlsx_path):
        """preview-slide returns base64 placeholder PNG if LibreOffice unavailable."""
        import shutil
        if shutil.which("soffice"):
            import pytest
            pytest.skip("LibreOffice is installed")

        state_dict = {
            "version": 1,
            "project_name": "Test",
            "inputs": {"db_path": str(valid_xlsx_path), "template_path": str(valid_template_path), "font_override": None},
            "slides": [],
        }

        req = {
            "state": state_dict,
            "slide_index": 0,
        }
        r = client.post("/api/preview-slide", json=req)
        assert r.status_code == 200
        body = r.json()
        assert "png_base64" in body
        assert isinstance(body["png_base64"], str)


class TestExportPptx:
    """Tests for /api/export-pptx endpoint"""

    def test_export_pptx_writes_file(self, valid_template_path, valid_xlsx_path, tmp_path):
        """export-pptx endpoint writes PPTX file to disk."""

        state_dict = {
            "version": 1,
            "project_name": "Export Test",
            "inputs": {"db_path": str(valid_xlsx_path), "template_path": str(valid_template_path), "font_override": None},
            "slides": [
                {
                    "id": "slide_0",
                    "type": "separator",
                    "title": "Export Test Slide",
                }
            ],
        }

        req = {
            "state": state_dict,
            "filename": "exported",
        }
        r = client.post("/api/export-pptx", json=req)
        assert r.status_code == 200
        assert r.headers["content-disposition"].startswith("attachment")
        assert "exported.pptx" in r.headers["content-disposition"]
        assert len(r.content) > 0


def test_export_pptx_is_download():
    state = {"project_name": "T", "inputs": {"db_path": "x", "template_path": "y"}}

    def fake_build(_state, path):
        from pathlib import Path
        Path(path).write_bytes(b"PPTXDATA")

    with patch("aurum_encuestas.api.build_pptx", side_effect=fake_build):
        r = client.post("/api/export-pptx", json={"state": state, "filename": "reporte"})
    assert r.status_code == 200
    cd = r.headers["content-disposition"]
    assert cd.startswith("attachment")
    assert "reporte.pptx" in cd
    assert r.content == b"PPTXDATA"


@patch("aurum_encuestas.api.generate_analysis")
def test_generate_analysis_endpoint(mock_gen):
    mock_gen.return_value = "El 50% respondió Sí."
    payload = {
        "scope": "chart",
        "context": {
            "section_title": "X", "question_text": "?", "options": ["Sí", "No"],
            "breakdown_label": "General",
            "data": {"Total": {"Sí": {"count": 50, "pct": 0.5}}},
        },
    }
    r = client.post("/api/generate-analysis", json=payload)
    assert r.status_code == 200
    assert "Sí" in r.json()["text"]


@patch("aurum_encuestas.api.generate_analysis")
def test_generate_analysis_returns_fallback_on_error(mock_gen):
    from aurum_encuestas.errors import LLMError
    mock_gen.side_effect = LLMError("API down")
    payload = {"scope": "chart", "context": {"section_title": "x", "question_text": "y", "options": [], "breakdown_label": "z", "data": {}}}
    r = client.post("/api/generate-analysis", json=payload)
    assert r.status_code == 200
    assert "[Análisis no disponible" in r.json()["text"]


# test_training_add_and_list, test_training_bank_returns_layouts, test_training_delete
# removed in M6.2 — M4/M5 training endpoints deleted; will be replaced by M6.8 corpus CRUD tests.

@patch("aurum_encuestas.api.suggest_layout")
def test_suggest_layout_endpoint(mock_sug):
    mock_sug.return_value = {"source": "ai", "elements": [{"role": "chart_0", "x": 100, "y": 100, "cx": 1000, "cy": 1000}]}
    payload = {
        "n_charts": 1, "chart_types": ["PIE"],
        "n_chart_an": 0, "n_q_an": 0, "has_slide_an": False,
        "free_area": {"x": 0, "y": 0, "cx": 12000000, "cy": 7000000},
    }
    r = client.post("/api/suggest-layout", json=payload)
    assert r.status_code == 200
    assert r.json()["source"] == "ai"


def test_uploads_are_session_isolated(valid_xlsx_path, tmp_path, monkeypatch):
    monkeypatch.setenv("HOME", str(tmp_path))
    with open(valid_xlsx_path, "rb") as f:
        data = f.read()
    files = {"file": ("data.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r1 = client.post("/api/parse-xlsx", files=files, headers={"X-Session-Id": "s1"})
    r2 = client.post("/api/parse-xlsx", files=files, headers={"X-Session-Id": "s2"})
    assert r1.status_code == 200 and r2.status_code == 200
    assert (tmp_path / ".aurum" / "sessions" / "s1" / "uploads" / "data.xlsx").exists()
    assert (tmp_path / ".aurum" / "sessions" / "s2" / "uploads" / "data.xlsx").exists()


def test_recents_add_and_list(tmp_path, monkeypatch):
    monkeypatch.setenv("HOME", str(tmp_path))
    # save a project (should auto-add to recents)
    proj = {"version": 1, "project_name": "P1", "inputs": {"db_path": "./x", "template_path": "./y", "font_override": None}, "slides": []}
    client.post("/api/save-project", json={"name": "p1", "state": proj}, headers={"X-Session-Id": "rec1"})

    r = client.get("/api/recents", headers={"X-Session-Id": "rec1"})
    assert r.status_code == 200
    recs = r.json()["recents"]
    assert any(rec["path"] == "p1" for rec in recs)


# ─── Task 1 (Fase AJ): /api/sheet-grid endpoint ──────────────────────────────


def test_sheet_grid_endpoint(valid_xlsx_path):
    r = client.post("/api/sheet-grid", json={"db_path": str(valid_xlsx_path)})
    assert r.status_code == 200
    body = r.json()
    assert body["n_rows"] > 0 and body["n_cols"] > 0
    assert body["truncated"] is False
    # Row 1 col D (index [0][3]) is the "Rango de edad" breakdown header
    assert body["cells"][0][3] == "Rango de edad"
    # Row 18 col A (index [17][0]) is the question marker
    assert body["cells"][17][0] == "$p1.recordacion"


def test_sheet_grid_bad_path():
    r = client.post("/api/sheet-grid", json={"db_path": "/no/such/file.xlsx"})
    assert r.status_code == 200
    assert "error" in r.json()


# ─── Fase AK T2: /api/cell-values endpoint ───────────────────────────────────


def _minimal_state(db, path: str) -> dict:
    """Return a ProjectState-compatible dict with parsed_db + inputs.db_path set."""
    from aurum_encuestas.models import ProjectState, ProjectInputs
    state = ProjectState(
        project_name="Test",
        inputs=ProjectInputs(db_path=path, template_path="/dev/null"),
        parsed_db=db,
    )
    return state.model_dump()


def test_cell_values_endpoint(valid_xlsx_path):
    from aurum_encuestas.xlsx_parser import parse_xlsx
    db = parse_xlsx(str(valid_xlsx_path))
    # Construct a valid ProjectState dict with parsed_db + inputs.db_path set to the fixture.
    state = _minimal_state(db, str(valid_xlsx_path))   # build per the ProjectState model
    r = client.post("/api/cell-values", json={
        "state": state, "question_id": db.questions[0].id, "breakdown_id": "general",
    })
    assert r.status_code == 200
    body = r.json()
    assert body["options"] == db.questions[0].options
    assert body["categories"] == ["Total"]
    assert body["cells"]["Sí"]["Total"]["count"] == 458


def test_count_cells_endpoint(valid_xlsx_path):
    from aurum_encuestas.xlsx_parser import parse_xlsx
    from aurum_encuestas.data_extractor import _find_question_rows, _resolve_breakdown_cols
    from openpyxl import load_workbook
    db = parse_xlsx(str(valid_xlsx_path))
    state = _minimal_state(db, str(valid_xlsx_path))
    r = client.post("/api/count-cells", json={"state": state})
    assert r.status_code == 200
    body = r.json()
    assert body.get("error") is None
    cells = body["cells"]
    assert len(cells) > 0
    # No duplicate coordinates
    pairs = [(c["row"], c["col"]) for c in cells]
    assert len(pairs) == len(set(pairs))
    # A known count cell (P1 "Sí" / general / Total) is present and holds 458
    ws = load_workbook(str(valid_xlsx_path), data_only=True).worksheets[0]
    q1 = db.questions[0]
    rows = _find_question_rows(ws, q1)
    cols = _resolve_breakdown_cols(ws, "general", db.data_blocks["counts_cols"][0])
    expected = {"row": rows["Sí"], "col": cols["Total"]}
    assert expected in cells
    assert ws.cell(expected["row"], expected["col"]).value == 458


def test_count_cells_endpoint_bad_state():
    r = client.post("/api/count-cells", json={"state": {"not": "valid"}})
    assert r.status_code == 200
    body = r.json()
    assert body["cells"] == []
    assert body["error"]


def test_suggest_slide_layout_positions_subtitle(monkeypatch, valid_xlsx_path):
    # Build a state with one shell slide containing a chart + a subtitle.
    from aurum_encuestas import api as api_mod
    state = {
        "project_name": "T",
        "inputs": {"db_path": str(valid_xlsx_path), "template_path": "y"},
        "parsed_db": None,
        "slides": [
            {"id": "s2", "type": "shell", "title": "Sec",
             "charts": [{"id": "c1", "question_id": "q1", "breakdown_ids": [], "chart_type": "PIE"}],
             "analyses": [],
             "subtitles": [{"id": "sub1", "text": "P1. Pregunta"}]},
        ],
    }
    # Stub the LLM so no network call happens; it returns a position for the subtitle.
    def fake_correct(slide_payload, slide_png_bytes=None, user_hint=None):
        # assert the subtitle shape was included in the payload
        kinds = [s.get("kind") for s in slide_payload["shapes"]]
        assert "subtitle" in kinds
        return {"elements": [{"id": "subtitle_sub1", "x_cm": 2.0, "y_cm": 3.0, "w_cm": 10.0, "h_cm": 1.0}]}
    monkeypatch.setattr(api_mod, "correct_slide_layout", fake_correct)
    # Avoid the PNG render path (build_pptx) failing in-test: monkeypatch render to None.
    monkeypatch.setattr(api_mod, "render_slide_to_png", lambda *a, **k: None)
    monkeypatch.setattr(api_mod, "build_pptx", lambda *a, **k: None)

    r = client.post("/api/suggest-slide-layout", json={"state": state, "slide_id": "s2", "user_hint": "x"})
    assert r.status_code == 200
    body = r.json()
    assert "sub1" in body["positions"]
    assert body["positions"]["sub1"]["cx_emu"] > 0


def test_suggest_slide_layout_propagates_box_style(monkeypatch, valid_xlsx_path):
    """box_style='dashed' from AI element reaches the response positions dict."""
    import aurum_encuestas.api as api
    monkeypatch.setattr(api, "correct_slide_layout", lambda *a, **k: {
        "elements": [
            {"id": "analysis_a1", "x_cm": 1.3, "y_cm": 4.0, "w_cm": 20.0, "h_cm": 3.0,
             "font_pt": 10.0, "box_style": "dashed"},
            {"id": "analysis_a2", "x_cm": 1.3, "y_cm": 8.0, "w_cm": 20.0, "h_cm": 3.0,
             "font_pt": 10.0, "box_style": "solid"},  # invalid → must collapse to None
        ],
        "extras": [], "changes": [],
    })
    from aurum_encuestas.models import ProjectState, ProjectInputs, Slide, Analysis
    state = ProjectState(
        project_name="Test",
        inputs=ProjectInputs(db_path=str(valid_xlsx_path), template_path="/dev/null"),
        slides=[Slide(
            id="slide1",
            type="shell",
            analyses=[
                Analysis(id="a1", scope="slide", text="Texto de análisis"),
                Analysis(id="a2", scope="slide", text="Otro análisis"),
            ],
        )],
    )
    r = client.post("/api/suggest-slide-layout", json={
        "state": state.model_dump(),
        "slide_id": "slide1",
    })
    assert r.status_code == 200
    body = r.json()
    assert "positions" in body
    assert "a1" in body["positions"], f"a1 not in positions: {list(body['positions'].keys())}"
    assert body["positions"]["a1"]["box_style"] == "dashed"
    # Invalid box_style value collapses to None (no arbitrary strings pass through).
    assert body["positions"]["a2"]["box_style"] is None
