import re
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .errors import XlsxParseError
from .models import Breakdown, ParsedDB, Question

BREAKDOWN_ID_MAP = {
    "rango de edad": "edad",
    "sexo": "sexo",
    "nse": "nse",
    "punto": "punto",
}

QMARKER_RE = re.compile(r"^\$p(\d+)\.(\w+)")


def _slug(text: str) -> str:
    return text.strip().lower()


def parse_xlsx(path: str) -> ParsedDB:
    try:
        wb = load_workbook(path, data_only=True)
    except (InvalidFileException, OSError, KeyError, BadZipFile) as e:
        raise XlsxParseError(f"No se pudo abrir el archivo: {e}") from e

    ws = wb.worksheets[0]

    breakdowns = _detect_breakdowns(ws)
    sample_size = _detect_sample_size(ws)
    questions = _detect_questions(ws)
    data_blocks = _detect_data_blocks(ws)

    return ParsedDB(
        questions=questions,
        breakdowns=breakdowns,
        sample_size=sample_size,
        data_blocks=data_blocks,
    )


def _detect_breakdowns(ws) -> list[Breakdown]:
    """Row 1 has breakdown group names in scattered cells. Row 2 has sub-categories below each group."""
    # Detect groups in row 1 (block 1 only — first occurrence)
    row1 = {c.column: (c.value or "") for c in ws[1]}
    row2 = {c.column: (c.value or "") for c in ws[2]}

    # Find "General" column (block 1 anchor)
    general_col = None
    for col, val in sorted(row2.items()):
        if str(val).strip() == "General":
            general_col = col
            break

    breakdowns = [Breakdown(id="general", label="General", categories=["Total"])]

    # Identify breakdown groups in row 1 (only block 1: cols 1 to general_col + range)
    block1_max = general_col + 30 if general_col else 50
    seen_labels = set()
    group_starts = []
    for col in sorted(row1.keys()):
        if col <= general_col or col > block1_max:
            continue
        label = str(row1[col]).strip()
        if not label or label in seen_labels:
            continue
        slug_key = _slug(label)
        if slug_key in BREAKDOWN_ID_MAP:
            seen_labels.add(label)
            group_starts.append((col, label, BREAKDOWN_ID_MAP[slug_key]))

    # For each group, categories are row2 cells from col to next group's col - 1
    sorted_groups = sorted(group_starts)
    for i, (col, label, gid) in enumerate(sorted_groups):
        end_col = sorted_groups[i + 1][0] if i + 1 < len(sorted_groups) else col + 6
        categories = []
        for c in range(col, end_col):
            v = row2.get(c)
            if v and str(v).strip() and str(v).strip() not in ("General",):
                categories.append(str(v).strip())
        if categories:
            breakdowns.append(Breakdown(id=gid, label=label, categories=categories))

    return breakdowns


def _detect_sample_size(ws) -> int:
    """Row 3 col 3 typically has Total = sample_size."""
    val = ws.cell(3, 3).value
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def _detect_questions(ws) -> list[Question]:
    """Scan col A for question markers ($pN.label or text ending in ?). Following rows = options."""
    questions = []
    current_q = None
    current_options = []
    next_qid = 1

    for row in range(3, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value

        if a_val is not None and str(a_val).strip():
            # New question row
            if current_q is not None and current_options:
                current_q.options = current_options
                questions.append(current_q)

            a_str = str(a_val).strip()
            m = QMARKER_RE.match(a_str)
            if m:
                code = f"P{m.group(1)}"
                text = m.group(0)
                confidence = 1.0
            elif a_str.endswith("?"):
                code = f"P{next_qid}"
                text = a_str
                confidence = 0.9
            else:
                # Could be a demographic row (Sexo, Rango de edad, NSE, Punto)
                if a_str.lower() in ("sexo", "rango de edad", "nse", "punto", "nse_a"):
                    current_q = None
                    current_options = []
                    continue
                # Treat as low-confidence question
                code = f"P{next_qid}"
                text = a_str
                confidence = 0.5

            next_qid += 1
            current_q = Question(id=f"q{next_qid - 1}", code=code, text=text, options=[], confidence=confidence)
            current_options = []
            if b_val is not None and str(b_val).strip():
                current_options.append(str(b_val).strip())
        elif current_q is not None and b_val is not None and str(b_val).strip():
            current_options.append(str(b_val).strip())

    if current_q is not None and current_options:
        current_q.options = current_options
        questions.append(current_q)

    return questions


def _detect_data_blocks(ws) -> dict:
    """Detect up to 3 column blocks by scanning value types in the first question row.

    A "counts" block has integers > 1. A "%" block has decimals 0-1.
    Blocks are separated by ≥1 empty/text column.

    Returns: {counts_cols: [start, end], pct_row_cols: [start, end], pct_col_cols: [start, end]}
    """
    # Find first question row (col A has a marker like $pN.label or text ending in ?)
    q_row = None
    for row in range(3, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        if a_val and (QMARKER_RE.match(str(a_val).strip()) or str(a_val).strip().endswith("?")):
            q_row = row
            break
    if q_row is None:
        return {"counts_cols": [3, 17], "pct_row_cols": [21, 35], "pct_col_cols": [41, 55]}

    # Classify each numeric column by value at q_row (or sum of next 2 rows if q_row is question header without data)
    def _classify(col: int) -> str:
        # try q_row, q_row+1, q_row+2 — take first numeric
        for r in (q_row, q_row + 1, q_row + 2):
            v = ws.cell(r, col).value
            if v is None or isinstance(v, str):
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            if 0 < fv <= 1.0:
                return "pct"
            if fv > 1:
                return "count"
        return "empty"

    # Walk columns from 3 onward, group contiguous same-type runs (allowing single empty gaps)
    max_col = ws.max_column
    blocks: list[tuple[int, int, str]] = []  # (start, end, kind)
    current_start = None
    current_kind = None
    prev_kind = "empty"
    for col in range(3, max_col + 1):
        k = _classify(col)
        if k == "empty":
            # close current block if any
            if current_start is not None:
                blocks.append((current_start, col - 1, current_kind))
                current_start = None
                current_kind = None
            prev_kind = "empty"
            continue
        if current_start is None:
            current_start = col
            current_kind = k
        elif k != current_kind:
            blocks.append((current_start, col - 1, current_kind))
            current_start = col
            current_kind = k
        prev_kind = k  # noqa: F841
    if current_start is not None:
        blocks.append((current_start, max_col, current_kind))

    # Pick the first count block + up to 2 pct blocks
    count_blocks = [b for b in blocks if b[2] == "count"]
    pct_blocks = [b for b in blocks if b[2] == "pct"]

    counts = (count_blocks[0][0], count_blocks[0][1]) if count_blocks else (3, 17)
    pct_row = (pct_blocks[0][0], pct_blocks[0][1]) if len(pct_blocks) >= 1 else (counts[1] + 4, counts[1] + 18)
    pct_col = (pct_blocks[1][0], pct_blocks[1][1]) if len(pct_blocks) >= 2 else (pct_row[1] + 4, pct_row[1] + 18)

    return {
        "counts_cols": [counts[0], counts[1]],
        "pct_row_cols": [pct_row[0], pct_row[1]],
        "pct_col_cols": [pct_col[0], pct_col[1]],
    }
