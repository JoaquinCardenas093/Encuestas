"""Build embedded xlsx for OLE TABLE_WITH_MINIBARS render."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Hex palette (no role mapping — direct hex to avoid color_resolver remapping)
HEADER_FILL_HEX = "FF999999"   # medium gray ARGB (alpha FF for full opacity)
HEADER_FONT_HEX = "FFFFFFFF"   # white ARGB
BODY_FILL_HEX = "FFFFFFFF"     # white ARGB
BODY_FONT_HEX = "FF000000"     # black ARGB
DATABAR_HEX = "FFD9D9D9"       # light gray bar ARGB
BORDER_HEX = "FF000000"        # cell borders black

HEADER_ROW = 2
CAT_ROW = 3
COUNTS_ROW = 4
FIRST_OPT_ROW = 5

LABEL_COL_W_MIN = 14
LABEL_COL_W_MAX = 40
DATA_COL_W = 14


# Excel column width unit ≈ 7px at 96 DPI for default font (Calibri 11).
# 1 px = 9525 EMU. So 1 col_width unit ≈ 66675 EMU.
EMU_PER_COL_W = 66675
EMU_PER_PT = 12700  # for row heights


def compute_xlsx_natural_dim_emu(source_chart, breakdown_groups: list[str]) -> tuple[int, int]:
    """Compute xlsx natural rendered dim (w_emu, h_emu) for given chart+bds.
    Uses default col widths from constants. Used to know "Excel would render this big"."""
    question = getattr(source_chart, "question", None)
    options = list(getattr(question, "options", []) or [])
    all_bds = getattr(source_chart, "all_breakdowns_data", {}) or {}
    bds = [(b, all_bds.get(b, {})) for b in breakdown_groups if b in all_bds]
    if not bds or not options:
        return 0, 0

    label_strings = list(options) + ["Observaciones"]
    longest = max((len(str(s)) for s in label_strings), default=0)
    label_col_w = max(LABEL_COL_W_MIN, min(LABEL_COL_W_MAX, longest + 2))

    total_cw = 0.0
    for i, (_, bd) in enumerate(bds):
        n_cats = len(bd.get("categories", {}) or {})
        if n_cats == 0:
            continue
        if i == 0:
            total_cw += label_col_w
        total_cw += n_cats * DATA_COL_W
        total_cw += 2  # spacer

    w_emu = int(total_cw * EMU_PER_COL_W)

    # rows: HEADER 24 + CAT 22 + COUNTS 18 + OPT 28 × N
    total_h_pt = 24 + 22 + 18 + 28 * len(options)
    h_emu = int(total_h_pt * EMU_PER_PT)
    return w_emu, h_emu


def build_xlsx_for_table(source_chart, breakdown_groups: list[str]) -> BytesIO:
    """Return in-memory xlsx with N independent tables side-by-side.

    One table per breakdown; each table has its own label col + group_header
    merge + cat sub-headers + counts row + option rows with DataBarRule.
    Spacer column between tables.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    # Hide worksheet gridlines (rely on explicit cell borders).
    ws.sheet_view.showGridLines = False
    # Page setup: landscape A3 + fit-to-width so libreoffice xlsx→PDF render
    # doesn't truncate wide tables.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.1

    question = getattr(source_chart, "question", None)
    options = list(getattr(question, "options", []) or [])
    all_bds = getattr(source_chart, "all_breakdowns_data", {}) or {}

    bds = [(bd_id, all_bds.get(bd_id, {})) for bd_id in breakdown_groups if bd_id in all_bds]

    show_legend = bool(getattr(source_chart, "show_legend", False))

    # Auto-size label col to fit longest option name + "Observaciones" header.
    # Excel col width unit ~ 1 char of Calibri 11. Add small padding.
    label_strings = list(options) + ["Observaciones"]
    longest = max((len(str(s)) for s in label_strings), default=0)
    label_col_w = max(LABEL_COL_W_MIN, min(LABEL_COL_W_MAX, longest + 2))

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    body_fill = PatternFill("solid", fgColor=BODY_FILL_HEX)
    header_font_bold_11 = Font(color=HEADER_FONT_HEX, bold=True, name="Calibri", size=11)
    header_font_bold_10 = Font(color=HEADER_FONT_HEX, bold=True, name="Calibri", size=10)
    body_font_10 = Font(color=BODY_FONT_HEX, name="Calibri", size=10)
    body_font_bold_11 = Font(color=BODY_FONT_HEX, bold=True, name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    _thin = Side(style="thin", color=BORDER_HEX)
    label_border = Border(bottom=_thin)  # Observaciones col: bottom only

    def _bdr(first: bool, last: bool, *, top=False, bottom=False, right_always=False):
        """Helper: build Border with side-set logic per row-type spec."""
        sides = {}
        if top:
            sides["top"] = _thin
        if bottom:
            sides["bottom"] = _thin
        if first:
            sides["left"] = _thin
        if last or right_always:
            sides["right"] = _thin
        return Border(**sides)

    cur_col = 1
    for bd_index, (bd_id, bd) in enumerate(bds):
        cats = bd.get("categories", {}) or {}
        n_cats = len(cats)
        if n_cats == 0:
            continue

        # Only the FIRST bd table renders the label col when show_legend=True.
        # Subsequent tables share the first table's leyenda visually (no own col).
        is_first_bd = (bd_index == 0)
        if show_legend and is_first_bd:
            label_col = cur_col
            data_start = cur_col + 1
        else:
            data_start = cur_col
            label_col = None
        data_end = data_start + n_cats - 1

        # Row 2: merged group_header (Sexo/NSE) — top + L + R, NO bottom.
        # NOTE: openpyxl merge_cells only applies master cell's border. Must set
        # borders on EACH cell in merge per position so outer top/right span fully.
        ws.merge_cells(
            start_row=HEADER_ROW, start_column=data_start,
            end_row=HEADER_ROW, end_column=data_end,
        )
        gh = ws.cell(row=HEADER_ROW, column=data_start,
                     value=bd.get("label") or bd_id)
        gh.fill = header_fill
        gh.font = header_font_bold_11
        gh.alignment = center
        # Apply per-cell borders across full merge span for continuous outer.
        for col_offset in range(n_cats):
            c = ws.cell(row=HEADER_ROW, column=data_start + col_offset)
            c.fill = header_fill  # ensure non-master cells also visually filled
            c.border = _bdr(
                first=(col_offset == 0),
                last=(col_offset == n_cats - 1),
                top=True,
            )

        # Row 3: cat sub-headers (Femenino/Masculino) — only bottom (+ L first / R last for outer).
        for i, (cat_label, _) in enumerate(cats.items()):
            ch = ws.cell(row=CAT_ROW, column=data_start + i, value=cat_label)
            ch.fill = header_fill
            ch.font = header_font_bold_10
            ch.alignment = center
            ch.border = _bdr(first=(i == 0), last=(i == n_cats - 1), bottom=True)

        # Row 4: counts row — label col = "Observaciones" (only on first bd), data cols = totals.
        # Counts cells: bottom only (no L/R divider between them).
        if label_col is not None:
            obs = ws.cell(row=COUNTS_ROW, column=label_col, value="Observaciones")
            obs.fill = body_fill
            obs.font = body_font_bold_11
            obs.alignment = right

        for i, (_, opt_cells) in enumerate(cats.items()):
            total = sum(int((opt_cells.get(o) or {}).get("count") or 0) for o in options)
            cc = ws.cell(row=COUNTS_ROW, column=data_start + i, value=total)
            cc.fill = body_fill
            cc.font = body_font_bold_11
            cc.alignment = center
            cc.border = _bdr(first=(i == 0), last=(i == n_cats - 1), bottom=True)

        # Rows 5+: option rows. Each option cell: right border always (between cols),
        # + left if first col (outer), + bottom if last option row (outer).
        n_opts = len(options)
        for j, opt in enumerate(options):
            row = FIRST_OPT_ROW + j

            if label_col is not None:
                lbl = ws.cell(row=row, column=label_col, value=opt)
                lbl.fill = body_fill
                lbl.font = body_font_bold_11
                lbl.alignment = right

            is_last_opt = (j == n_opts - 1)
            for i, (_, opt_cells) in enumerate(cats.items()):
                pct = float((opt_cells.get(opt) or {}).get("pct") or 0)
                oc = ws.cell(row=row, column=data_start + i, value=pct)
                oc.number_format = "0.0%"
                oc.fill = body_fill
                oc.font = body_font_10
                oc.alignment = Alignment(horizontal="left", indent=1, vertical="center")
                oc.border = _bdr(
                    first=(i == 0),
                    last=(i == n_cats - 1),
                    bottom=is_last_opt,
                    right_always=True,
                )

        # DataBarRule per OPTION ROW spanning this bd's data cols only
        for j in range(len(options)):
            row = FIRST_OPT_ROW + j
            start_letter = get_column_letter(data_start)
            end_letter = get_column_letter(data_end)
            range_str = f"{start_letter}{row}:{end_letter}{row}"
            rule = DataBarRule(
                start_type="num", start_value=0,
                end_type="num", end_value=1,
                color=DATABAR_HEX,
                showValue=True,
            )
            # gradient=False matches target design (solid bar fill).
            if rule.dataBar is not None:
                rule.dataBar.gradient = False
            ws.conditional_formatting.add(range_str, rule)

        # Column widths for this bd
        if label_col is not None:
            ws.column_dimensions[get_column_letter(label_col)].width = label_col_w
        for c in range(data_start, data_end + 1):
            ws.column_dimensions[get_column_letter(c)].width = DATA_COL_W

        # Explicit narrow width for spacer col between bds
        ws.column_dimensions[get_column_letter(data_end + 1)].width = 2
        cur_col = data_end + 2

    ws.row_dimensions[HEADER_ROW].height = 24
    ws.row_dimensions[CAT_ROW].height = 22
    ws.row_dimensions[COUNTS_ROW].height = 18
    for j in range(len(options)):
        ws.row_dimensions[FIRST_OPT_ROW + j].height = 28

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _force_databar_solid(buf)


# ---- x14:dataBar gradient="0" injection (Fase Q) ----
# Real Office Excel-embed xlsx files (verified in PPT Aurora ejemplo.pptx
# Microsoft_Excel_Worksheet4.xlsx) use this exact structure for solid bars.
# Previous attempts failed because of malformed GUIDs (template "% seq"
# generated 13-char last segment) + missing required negativeFillColor /
# axisColor children. Excel surfaces this as "Memoria insuficiente para
# leer Hoja de cálculo" when the OLE Excel host validates the embed.

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_X14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_NS_XM = "http://schemas.microsoft.com/office/excel/2006/main"
_EXT_URI_CF_RULE = "{B025F937-C7B1-47D3-B67F-A62EFF666E3E}"
_EXT_URI_CF_LIST = "{78C0D931-6437-407d-A8EE-F0AAD7539E65}"  # lowercase 'd'


def _new_guid() -> str:
    import uuid
    return "{" + str(uuid.uuid4()).upper() + "}"


def _force_databar_solid(xlsx_buf: BytesIO) -> BytesIO:
    """Post-process xlsx to add x14:dataBar gradient="0" override."""
    import zipfile

    xlsx_buf.seek(0)
    out = BytesIO()
    with zipfile.ZipFile(xlsx_buf, "r") as zin:
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                    try:
                        data = _inject_x14_solid_databar(data)
                    except Exception:
                        pass  # Fall back to gradient bars on any error.
                zout.writestr(item, data)
    out.seek(0)
    return out


def _inject_x14_solid_databar(sheet_bytes: bytes) -> bytes:
    """Inject x14:dataBar gradient="0" override matching real Office output."""
    from lxml import etree

    parser = etree.XMLParser(remove_blank_text=False)
    root = etree.fromstring(sheet_bytes, parser=parser)
    ns = {"main": _NS_MAIN}

    cf_blocks = root.findall("main:conditionalFormatting", ns)
    if not cf_blocks:
        return sheet_bytes

    pairs: list[tuple[str, str, list]] = []  # (sqref, guid, inner_cfvos)
    for cf in cf_blocks:
        sqref = cf.get("sqref")
        for rule in cf.findall('main:cfRule[@type="dataBar"]', ns):
            inner_db = rule.find("main:dataBar", ns)
            if inner_db is None:
                continue
            inner_cfvos = inner_db.findall("main:cfvo", ns)
            guid = _new_guid()

            # Append <extLst><ext uri="..." xmlns:x14="..."><x14:id>GUID</x14:id></ext></extLst>
            # AFTER the inner <dataBar> (per real Office layout).
            extlst = etree.SubElement(rule, f"{{{_NS_MAIN}}}extLst")
            ext = etree.SubElement(
                extlst,
                f"{{{_NS_MAIN}}}ext",
                attrib={"uri": _EXT_URI_CF_RULE},
                nsmap={"x14": _NS_X14},
            )
            x14_id = etree.SubElement(ext, f"{{{_NS_X14}}}id")
            x14_id.text = guid
            pairs.append((sqref, guid, list(inner_cfvos)))

    if not pairs:
        return sheet_bytes

    # Append to existing worksheet-level extLst, OR create one.
    extlst = root.find("main:extLst", ns)
    if extlst is None:
        extlst = etree.SubElement(root, f"{{{_NS_MAIN}}}extLst")

    ext = etree.SubElement(
        extlst,
        f"{{{_NS_MAIN}}}ext",
        attrib={"uri": _EXT_URI_CF_LIST},
        nsmap={"x14": _NS_X14},
    )
    x14_cfs = etree.SubElement(ext, f"{{{_NS_X14}}}conditionalFormattings")

    for sqref, guid, inner_cfvos in pairs:
        x14_cf = etree.SubElement(
            x14_cfs,
            f"{{{_NS_X14}}}conditionalFormatting",
            nsmap={"xm": _NS_XM},
        )
        x14_rule = etree.SubElement(
            x14_cf,
            f"{{{_NS_X14}}}cfRule",
            attrib={"type": "dataBar", "id": guid},
        )
        x14_databar = etree.SubElement(
            x14_rule,
            f"{{{_NS_X14}}}dataBar",
            attrib={"minLength": "0", "maxLength": "100", "gradient": "0"},
        )
        # Mirror inner cfvo types to ext.
        for c in inner_cfvos:
            kind = c.get("type")
            if kind == "num":
                cfvo = etree.SubElement(
                    x14_databar, f"{{{_NS_X14}}}cfvo", attrib={"type": "num"}
                )
                f_el = etree.SubElement(cfvo, f"{{{_NS_XM}}}f")
                f_el.text = c.get("val") or "0"
            elif kind == "min":
                etree.SubElement(
                    x14_databar, f"{{{_NS_X14}}}cfvo", attrib={"type": "autoMin"}
                )
            elif kind == "max":
                etree.SubElement(
                    x14_databar, f"{{{_NS_X14}}}cfvo", attrib={"type": "autoMax"}
                )
            else:
                etree.SubElement(
                    x14_databar, f"{{{_NS_X14}}}cfvo", attrib={"type": "autoMin"}
                )
        # REQUIRED children per real Office schema.
        etree.SubElement(
            x14_databar,
            f"{{{_NS_X14}}}negativeFillColor",
            attrib={"rgb": "FFFF0000"},
        )
        etree.SubElement(
            x14_databar,
            f"{{{_NS_X14}}}axisColor",
            attrib={"rgb": "FF000000"},
        )

        xm_sqref = etree.SubElement(x14_cf, f"{{{_NS_XM}}}sqref")
        xm_sqref.text = sqref

    return etree.tostring(
        root, xml_declaration=True, encoding="UTF-8", standalone=True
    )
