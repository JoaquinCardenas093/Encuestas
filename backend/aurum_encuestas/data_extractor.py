from openpyxl import load_workbook

from .models import Breakdown, Question
from .xlsx_parser import _slug, BREAKDOWN_ID_MAP


def _override_key(question_id: str, breakdown_id: str, category: str, option: str) -> str:
    return f"{question_id}|{breakdown_id}|{category}|{option}"


def _count_cell_set(count_cells: list | None) -> set[tuple[int, int]] | None:
    """list[[row,col]] → set of (row,col); None/empty → None (no filter)."""
    if not count_cells:
        return None
    return {(int(r), int(c)) for r, c in count_cells}


def extract_chart_data(xlsx_path: str, question: Question, breakdown_id: str,
                       data_blocks: dict, allowed_categories: list[str] | None = None,
                       total_row: int | None = None,
                       overrides: dict | None = None,
                       count_cells: list | None = None) -> dict:
    """Returns {breakdown_category: {option: {count, pct}}}.
    pct = count / column_total, where column_total = cell(total_row, col)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    q_rows = _find_question_rows(ws, question)
    counts_start = data_blocks["counts_cols"][0]
    breakdown_cols = _resolve_breakdown_cols(ws, breakdown_id, counts_start)
    if allowed_categories is not None:
        breakdown_cols = {c: breakdown_cols[c] for c in allowed_categories if c in breakdown_cols}

    cset = _count_cell_set(count_cells)
    result: dict[str, dict[str, dict]] = {}
    for cat, col in breakdown_cols.items():
        result[cat] = {}
        total = ws.cell(total_row, col).value if total_row else None
        try:
            total_v = float(total) if total is not None else None
        except (TypeError, ValueError):
            total_v = None
        for opt, row in q_rows.items():
            count = ws.cell(row, col).value or 0
            try:
                count_v = int(count)
            except (TypeError, ValueError):
                count_v = 0
            if cset is not None and (row, col) not in cset:
                count_v = 0
            pct_v = (count_v / total_v) if (total_v and total_v != 0) else None
            if overrides:
                ov = overrides.get(_override_key(getattr(question, "id", ""), breakdown_id, cat, opt))
                if ov:
                    if ov.get("count") is not None:
                        count_v = ov["count"]
                    if ov.get("pct") is not None:
                        pct_v = ov["pct"]
            result[cat][opt] = {"count": count_v, "pct": pct_v}
    return result


def extract_all_breakdowns_data(
    xlsx_path: str,
    question: Question,
    breakdowns: list[Breakdown],
    data_blocks: dict,
    total_row: int | None = None,
    overrides: dict | None = None,
    count_cells: list | None = None,
) -> dict:
    """Returns nested structure with every breakdown group for the question.

    Shape:
        {
            group_id: {
                "label": str,                                   # display name
                "categories": {
                    category_label: {
                        option_text: {"count": int, "pct": float|None}
                    }
                }
            }
        }

    Used by segmented-table renderer to build group-merged headers without
    needing N separate chart_ref_index queries.
    pct = count / column_total, where column_total = cell(total_row, col).
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    q_rows = _find_question_rows(ws, question)
    counts_start = data_blocks["counts_cols"][0]
    cset = _count_cell_set(count_cells)

    result: dict[str, dict] = {}
    for bd in breakdowns:
        cnt_cols = _resolve_breakdown_cols(ws, bd.id, counts_start)
        cats: dict[str, dict] = {}
        for cat in bd.categories:  # preserve declared order
            col = cnt_cols.get(cat)
            if col is None:
                continue
            total = ws.cell(total_row, col).value if total_row else None
            try:
                total_v = float(total) if total is not None else None
            except (TypeError, ValueError):
                total_v = None
            cell_map: dict[str, dict] = {}
            for opt, row in q_rows.items():
                count = ws.cell(row, col).value or 0
                try:
                    count_v = int(count)
                except (TypeError, ValueError):
                    count_v = 0
                if cset is not None and (row, col) not in cset:
                    count_v = 0
                pct_v = (count_v / total_v) if (total_v and total_v != 0) else None
                if overrides:
                    ov = overrides.get(_override_key(getattr(question, "id", ""), bd.id, cat, opt))
                    if ov:
                        if ov.get("count") is not None:
                            count_v = ov["count"]
                        if ov.get("pct") is not None:
                            pct_v = ov["pct"]
                cell_map[opt] = {"count": count_v, "pct": pct_v}
            cats[cat] = cell_map
        result[bd.id] = {"label": bd.label, "categories": cats}
    return result


def _find_question_rows(ws, question: Question) -> dict[str, int]:
    """Find rows for this question's options.

    Primary: anchor on the col-A marker (== question.text or a variable-code
    prefix like '$p3.llamado' for the sheet's '$p3.llamado.acción'), then collect
    col-B options under it. Fallback (multi-response questions whose col-A marker
    doesn't match question.text at all): scan col B for the option texts directly.
    """
    q_text = (question.text or "").strip()
    rows: dict[str, int] = {}
    in_question = False
    options_left = list(question.options)

    def _is_marker(a_s: str) -> bool:
        # Exact, or one is a dotted-prefix of the other ($p3.llamado vs $p3.llamado.acción).
        return bool(a_s) and (
            a_s == q_text
            or a_s.startswith(q_text + ".")
            or q_text.startswith(a_s + ".")
        )

    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""
        if _is_marker(a_s):
            in_question = True
            if b_s in options_left:
                rows[b_s] = r
                options_left.remove(b_s)
        elif in_question and not a_s and b_s in options_left:
            rows[b_s] = r
            options_left.remove(b_s)
        elif in_question and a_s:
            break  # next question started
    if rows:
        return rows

    # Fallback: match purely by option text in col B (first occurrence). Option
    # strings are long, unique sentences so cross-question collisions are unlikely.
    opts = {str(o).strip() for o in question.options}
    for r in range(3, ws.max_row + 1):
        b = ws.cell(r, 2).value
        b_s = str(b).strip() if b is not None else ""
        if b_s in opts and b_s not in rows:
            rows[b_s] = r
    return rows


def _resolve_breakdown_cols(ws, breakdown_id: str, block_start_col: int) -> dict[str, int]:
    """Map category label → column for the given breakdown within the column block.

    Generic: find the row-1 header at/after block_start_col whose slug equals
    breakdown_id, or whose alias (BREAKDOWN_ID_MAP) equals breakdown_id.
    """
    row2 = {c.column: (c.value or "") for c in ws[2]}

    if breakdown_id == "general":
        return {"Total": block_start_col}

    row1 = {c.column: (c.value or "") for c in ws[1]}
    sorted_cols = sorted(c for c in row1.keys() if c >= block_start_col)
    group_starts = [(c, str(row1[c]).strip()) for c in sorted_cols if str(row1[c]).strip()]

    found = None
    for i, (c, label) in enumerate(group_starts):
        slug = _slug(label)
        if slug == breakdown_id or BREAKDOWN_ID_MAP.get(slug) == breakdown_id:
            end_col = group_starts[i + 1][0] if i + 1 < len(group_starts) else c + 7
            found = (c, end_col)
            break

    if not found:
        return {}

    start, end = found
    out: dict[str, int] = {}
    for c in range(start, end):
        cat = str(row2.get(c) or "").strip()
        if cat:
            out[cat] = c
    return out
